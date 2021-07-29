# -*- coding: utf-8 -*-
"""
Created on Mon Jan 18 21:33:10 2021

@author: Devineni
"""

import pandas as pd
import numpy as np
from statistics import mean
import time
import datetime as dt
import matplotlib.pyplot as plt
from tabulate import tabulate

# import mysql.connector
import os
import pymysql
from sqlalchemy import create_engine

from openpyxl import load_workbook
import statistics
from easygui import *
import sys

def prRed(skk): print("\033[31;1;m {}\033[00m" .format(skk)) 


from uncertainties import ufloat
engine = create_engine("mysql+pymysql://root:Password123@localhost/",pool_pre_ping=True)

#%%%
import datetime
import matplotlib.dates as mdates

import matplotlib.units as munits
from pylab import rcParams
rcParams['figure.figsize'] = 7,4.5
# rcParams['figure.figsize'] = 19,15 # word

plt.rcParams["font.family"] = "calibri"
plt.rcParams["font.weight"] = "normal"
plt.rcParams["font.size"] = 10



#%% Automated Outdoor results
'''
This section deals with taking input selection of the experiment
easygui module was used to create the dialogue boxes for easy input
this is just a more visual way for experiment selection
'''
i=0
result1 = pd.DataFrame(["Outdoor Summary"])
databases = ["ESHL_summer", "ESHL_winter", "CBo_summer", "CBo_winter"]
for database in databases:
    times = pd.read_excel('C:/Users/Raghavakrishna/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/excel_files/Times_thesis.xlsx', sheet_name= database)

    choices = list(times['short name'])
    for experiment in choices:
        z = int(times[times['short name'] == experiment].index.values)
        Vdot_sheets = {"ESHL_summer":"ESHL_Vdot", "ESHL_winter":"ESHL_Vdot", "CBo_summer":"CBo_Vdot", "CBo_winter":"CBo_Vdot"}
        
        
        t0 = times.loc[z,"Start"]
        tn = times.loc[z,"End"]
        #%%%
        result = pd.DataFrame(["Outdoor Results"]) ;result.loc[0,1] = experiment
        r = 1
        
        adf = pd.read_sql_query("SELECT * FROM weather.außen WHERE datetime BETWEEN '{}' AND '{}'".format(t0,tn), con = engine).drop("index", axis = 1).set_index("datetime")
        result.loc[r,0] = "parameter" ; result.loc[r,1] = "min" ; result.loc[r,2] = "max"; result.loc[r,3] = "mean" ; result.loc[r,4] = "std"
        r += 1
        result.loc[r,0] = "temp_°C" ; result.loc[r,1] = adf["temp_°C"].min() ; result.loc[r,2] = adf["temp_°C"].max(); result.loc[r,3] = adf["temp_°C"].mean() ; result.loc[r,4] = adf["temp_°C"].std()
        r += 1
        result.loc[r,0] = "RH_%rH" ; result.loc[r,1] = adf["RH_%rH"].min() ; result.loc[r,2] = adf["RH_%rH"].max(); result.loc[r,3] = adf["RH_%rH"].mean() ; result.loc[r,4] = adf["RH_%rH"].std()
        r += 1
        result.loc[r,0] = "CO2_ppm" ; result.loc[r,1] = adf["CO2_ppm"].min() ; result.loc[r,2] = adf["CO2_ppm"].max(); result.loc[r,3] = adf["CO2_ppm"].mean() ; result.loc[r,4] = adf["CO2_ppm"].std()
        
        wdf = pd.read_sql_query("SELECT * FROM weather.weather_all WHERE datetime BETWEEN '{}' AND '{}'".format(t0,tn), con = engine).set_index("datetime")
        r += 1
        result.loc[r,0] = "Wind Speed, m/s" ; result.loc[r,1] = wdf["Wind Speed, m/s"].min() ; result.loc[r,2] = wdf["Wind Speed, m/s"].max(); result.loc[r,3] = wdf["Wind Speed, m/s"].mean() ; result.loc[r,4] = wdf["Wind Speed, m/s"].std()
        r += 1
        result.loc[r,0] = "Gust Speed, m/s" ; result.loc[r,1] = wdf["Gust Speed, m/s"].min() ; result.loc[r,2] = wdf["Gust Speed, m/s"].max(); result.loc[r,3] = wdf["Gust Speed, m/s"].mean() ; result.loc[r,4] = wdf["Gust Speed, m/s"].std()
        
        
        
        #%%%writer
        # path = "C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/results/outdoor_results.xlsx"
        
        # book = load_workbook(path)
        # writer = pd.ExcelWriter(path, engine = 'openpyxl')
        # writer.book = book
        
        # result.to_excel(writer, index = False , sheet_name = experiment[:10])
        # writer.save()
        # writer.close()
        #%%%
        result1.loc[i+1,0] = experiment; result1.loc[i+1,1] = adf["temp_°C"].mean(); result1.loc[i+1,2] = adf["temp_°C"].std()
        result1.loc[i+1,3] = adf["RH_%rH"].mean();result1.loc[i+1,4] = adf["RH_%rH"].std()
        result1.loc[i+1,5] = adf["CO2_ppm"].mean();result1.loc[i+1,6] = adf["CO2_ppm"].std()
        result1.loc[i+1,7] = wdf["Wind Speed, m/s"].mean();result1.loc[i+1,8] = wdf["Wind Speed, m/s"].std()
        result1.loc[i+1,9] = wdf["Gust Speed, m/s"].mean();result1.loc[i+1,10] = wdf["Gust Speed, m/s"].std();
        result1.loc[i+1,11] = wdf["Wind Direction"].mean();result1.loc[i+1,12] = wdf["Wind Direction"].std()
        result1.loc[i+1,13] = wdf["Temperature °C"].mean();result1.loc[i+1,14] = wdf["Temperature °C"].std();
        result1.loc[i+1,15] = wdf["RH %"].mean();result1.loc[i+1,16] = wdf["RH %"].std()
        i = i+1
        
result1.columns = ["experiment", "temp_°C","std","RH_%rH","std","CO2_ppm" ,"std","Wind Speed, m/s","std","Gust Speed, m/s","std","Wind Direction","std", "Temperature °C","std", "RH %","std"]

#%%%writer
# path = "C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/results/outdoor_results.xlsx"

# book = load_workbook(path)
# writer = pd.ExcelWriter(path, engine = 'openpyxl')
# writer.book = book

# result1.to_excel(writer, index = False , sheet_name = "summary")
# writer.save()
# writer.close()
#%%% Manual selection
# '''
# This section deals with taking input selection of the experiment
# easygui module was used to create the dialogue boxes for easy input
# this is just a more visual way for experiment selection
# '''

# msg ="Please select a Location/Season you like to analyze"
# title = "Season selection"
# choices = ["ESHL_summer", "ESHL_winter", "CBo_summer", "CBo_winter"]
# database = choicebox(msg, title, choices)


# times = pd.read_excel('C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/excel_files/Times_thesis.xlsx', sheet_name= database)

# msg ="Please select an experiment you would like to analyse in {database}".format(database = str(database))
# title = "Experiment selection"
# choices = list(times['short name'])
# experiment = choicebox(msg, title, choices)


# z = int(times[times['short name'] == experiment].index.values)
# Vdot_sheets = {"ESHL_summer":"ESHL_Vdot", "ESHL_winter":"ESHL_Vdot", "CBo_summer":"CBo_Vdot", "CBo_winter":"CBo_Vdot"}


# t0 = times.loc[z,"Start"]
# tn = times.loc[z,"End"]
# #%%
# result = pd.DataFrame(["Outdoor Results"]) ;result.loc[0,1] = experiment
# r = 1

# adf = pd.read_sql_query("SELECT * FROM weather.außen WHERE datetime BETWEEN '{}' AND '{}'".format(t0,tn), con = engine).drop("index", axis = 1).set_index("datetime")
# result.loc[r,0] = "parameter" ; result.loc[r,1] = "min" ; result.loc[r,2] = "max"; result.loc[r,3] = "mean" ; result.loc[r,4] = "std"
# r += 1
# result.loc[r,0] = "temp_°C" ; result.loc[r,1] = adf["temp_°C"].min() ; result.loc[r,2] = adf["temp_°C"].max(); result.loc[r,3] = adf["temp_°C"].mean() ; result.loc[r,4] = adf["temp_°C"].std()
# r += 1
# result.loc[r,0] = "RH_%rH" ; result.loc[r,1] = adf["RH_%rH"].min() ; result.loc[r,2] = adf["RH_%rH"].max(); result.loc[r,3] = adf["RH_%rH"].mean() ; result.loc[r,4] = adf["RH_%rH"].std()
# r += 1
# result.loc[r,0] = "CO2_ppm" ; result.loc[r,1] = adf["CO2_ppm"].min() ; result.loc[r,2] = adf["CO2_ppm"].max(); result.loc[r,3] = adf["CO2_ppm"].mean() ; result.loc[r,4] = adf["CO2_ppm"].std()

# wdf = pd.read_sql_query("SELECT * FROM weather.weather_all WHERE datetime BETWEEN '{}' AND '{}'".format(t0,tn), con = engine).set_index("datetime")
# r += 1
# result.loc[r,0] = "Wind Speed, m/s" ; result.loc[r,1] = wdf["Wind Speed, m/s"].min() ; result.loc[r,2] = wdf["Wind Speed, m/s"].max(); result.loc[r,3] = wdf["Wind Speed, m/s"].mean() ; result.loc[r,4] = wdf["Wind Speed, m/s"].std()
# r += 1
# result.loc[r,0] = "Gust Speed, m/s" ; result.loc[r,1] = wdf["Gust Speed, m/s"].min() ; result.loc[r,2] = wdf["Gust Speed, m/s"].max(); result.loc[r,3] = wdf["Gust Speed, m/s"].mean() ; result.loc[r,4] = wdf["Gust Speed, m/s"].std()



# #%%
# path = "C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/results/outdoor_results.xlsx"

# book = load_workbook(path)
# writer = pd.ExcelWriter(path, engine = 'openpyxl')
# writer.book = book

# result.to_excel(writer, index = False , sheet_name = experiment[:10])
# writer.save()
# writer.close()

#%% Wall temperature Indoor
i=0
wall_dict = {"ESHL_summer":"eshl_summer_wall", "ESHL_winter":"eshl_winter_wall", "CBo_summer":"cbo_summer_wall", "CBo_winter":"cbo_winter_wall"}

result_wall = pd.DataFrame(["Wall Summary"])
databases = ["ESHL_summer", "ESHL_winter", "CBo_summer", "CBo_winter"]
for database in databases:
    times = pd.read_excel('C:/Users/Raghavakrishna/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/excel_files/Times_thesis.xlsx', sheet_name= database)

    choices = list(times['short name'])
    for experiment in choices:
        z = int(times[times['short name'] == experiment].index.values)
        Vdot_sheets = {"ESHL_summer":"ESHL_Vdot", "ESHL_winter":"ESHL_Vdot", "CBo_summer":"CBo_Vdot", "CBo_winter":"CBo_Vdot"}
        
        
        t0 = times.loc[z,"Start"]
        tn = times.loc[z,"End"]

#%%
        schema = "weather"
        ''' this engine is used where ever connection is required to database'''
        engine = create_engine("mysql+pymysql://root:Password123@localhost/{}".format(schema),pool_pre_ping=True)
        
        wadf = pd.read_sql_query("SELECT * FROM weather.{} WHERE datetime BETWEEN '{}' AND '{}'".format(wall_dict[database],t0,tn), con = engine).set_index("datetime")
        
        
        result_wall.loc[i+1,0] = experiment; result_wall.loc[i+1,1] = wadf.mean().mean();result_wall.loc[i+1,2] = wadf.mean().std()
        i = i+1

result_wall.columns = ["experiment", "temp_°C", "std"] 
#%%writer
# path = "C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/results/indoor_results.xlsx"

# book = load_workbook(path)
# writer = pd.ExcelWriter(path, engine = 'openpyxl')
# writer.book = book

# result_wall.to_excel(writer, index = False , sheet_name = "wall_temp")
# writer.save()
# writer.close()



#%% Humidity Indoor
result_humidity = pd.DataFrame(["Humidity Summary"])
schema = "weather"
''' this engine is used where ever connection is required to database'''
engine = create_engine("mysql+pymysql://root:Password123@localhost/{}".format(schema),pool_pre_ping=True)
 
i=0
result1 = pd.DataFrame(["Outdoor Summary"])
databases = ["ESHL_summer", "ESHL_winter", "CBo_summer", "CBo_winter"]
for database in databases:
    times = pd.read_excel('C:/Users/Raghavakrishna/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/excel_files/Times_thesis.xlsx', sheet_name= database)

    choices = list(times['short name'])
    for experiment in choices:
        z = int(times[times['short name'] == experiment].index.values)
        Vdot_sheets = {"ESHL_summer":"ESHL_Vdot", "ESHL_winter":"ESHL_Vdot", "CBo_summer":"CBo_Vdot", "CBo_winter":"CBo_Vdot"}
        
        
        t0 = times.loc[z,"Start"]
        tn = times.loc[z,"End"]
        
        names = pd.read_sql_query('SHOW TABLES FROM {}'.format(database.lower()), engine)


        exclude_always = ['bd_original', 'außen', 'weather', 'tr']
    
        exclude = exclude_always + times["exclude"][z].split(',')
    
        tables = names[~names["Tables_in_{}".format(database.lower())].isin(exclude)].iloc[:,0]
        humidity = []
        
        for table in tables:
            hudf = pd.read_sql_query("SELECT * FROM {}.{} WHERE datetime BETWEEN '{}' AND '{}'".format(database.lower(),table,t0,tn), con = engine).set_index("datetime").dropna()
            
            if 'RH_%rH' in hudf.columns:
                humidity.append(hudf["RH_%rH"].mean())
        humidity = [x for x in humidity if x == x]        
        print(humidity)       
        result_humidity.loc[i,0] = experiment; result_humidity.loc[i,1] = mean(humidity); result_humidity.loc[i,2] = statistics.stdev(humidity)
        i = i+1


result_humidity = result_humidity.round(0)
#%%
# path = "C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/results/indoor_results.xlsx"

# book = load_workbook(path)
# writer = pd.ExcelWriter(path, engine = 'openpyxl')
# writer.book = book

# result_humidity.to_excel(writer, index = False , sheet_name = "humidity")
# writer.save()
# writer.close()


#%%

#%% wind speeds indoor

r = 1
result = pd.DataFrame(["Velocity Results in m/sec"]) ;result.loc[0,1] = "velocity m/sec";result.loc[0,2] = "std";result.loc[0,3] = "min";result.loc[0,4] = "max"
databases = [ "CBo_summer", "CBo_winter", "ESHL_summer", "ESHL_winter"]
for database in databases:
    times = pd.read_excel('C:/Users/Raghavakrishna/OneDrive - bwedu/MA_Raghavakrishna/1_Evaluation/excel_files/Times_thesis.xlsx', sheet_name= database)

    choices = list(times['short name'])
    for experiment in choices:
        z = int(times[times['short name'] == experiment].index.values)
        t0 = times.loc[z,"Start"]
        tn = times.loc[z,"End"]
        tables = ["1a_testo","2a_testo","3a_testo","4a_testo"]
        for table in tables:
            sdf = pd.read_sql_query("SELECT * FROM {}.{} WHERE datetime BETWEEN '{}' AND '{}'".format(database.lower(),table,t0,tn), con = engine)

            sdf = sdf.drop_duplicates(subset="datetime").set_index("datetime")
            sdf = sdf.loc[:,["hw_m/sec"]].dropna()

            result.loc[r,0] = table; result.loc[r,1] = sdf["hw_m/sec"].mean(); result.loc[r,2] = sdf["hw_m/sec"].std(); result.loc[r,3] = sdf["hw_m/sec"].max();result.loc[r,4] = sdf["hw_m/sec"].min()
            result.loc[r,5] = experiment
            r = r+1


#%%% writer
# path = "C:/Users/Raghavakrishna/OneDrive - bwedu/MA_Raghavakrishna/1_Evaluation/results/indoor_results.xlsx"

# book = load_workbook(path)
# writer = pd.ExcelWriter(path, engine = 'openpyxl')
# writer.book = book

# result.to_excel(writer, index = False , sheet_name = "velocity")
# writer.save()
# writer.close()








